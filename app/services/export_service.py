from __future__ import annotations

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models import AuditEvent, CalculationRun, CalculationSnapshot, CalculationStatus
from app.utils import canonical_json


class ExportService:
    def __init__(self, db: Session):
        self.db = db

    def _load(self, run_id: str) -> CalculationRun:
        run = self.db.scalar(
            select(CalculationRun)
            .options(selectinload(CalculationRun.summary), selectinload(CalculationRun.contributions))
            .where(CalculationRun.id == run_id)
        )
        if not run:
            raise ValueError("Calculation not found")
        if run.status != CalculationStatus.approved:
            raise PermissionError("Only approved results can be formally exported")
        return run

    def as_json(self, run_id: str) -> bytes:
        run = self._load(run_id)
        snapshot = self.db.get(CalculationSnapshot, run.snapshot_id)
        events = list(
            self.db.scalars(
                select(AuditEvent)
                .where(AuditEvent.object_id == run.id)
                .order_by(AuditEvent.occurred_at)
            )
        )
        summary = run.summary
        payload = {
            "calculation_id": run.id,
            "status": run.status.value,
            "manifest_hash": run.manifest_hash,
            "snapshot_manifest_hash": snapshot.manifest_hash,
            "summary": {
                column.name: getattr(summary, column.name)
                for column in summary.__table__.columns
                if column.name not in {"id", "run_id", "created_at"}
            },
            "contributions": [
                {
                    "dimension": x.dimension,
                    "code": x.code,
                    "name": x.name,
                    "amount_kg_co2e": x.amount_kg_co2e,
                    "rank": x.rank,
                    "metadata": x.metadata_json,
                }
                for x in run.contributions
            ],
            "audit_events": [
                {
                    "occurred_at": e.occurred_at,
                    "actor": e.actor,
                    "action": e.action,
                    "details": e.details,
                }
                for e in events
            ],
        }
        return canonical_json(payload)

    def as_excel(self, run_id: str) -> bytes:
        run = self._load(run_id)
        snapshot = self.db.get(CalculationSnapshot, run.snapshot_id)
        wb = Workbook()
        ws = wb.active
        ws.title = "摘要"
        blue = PatternFill("solid", fgColor="17365D")
        white_bold = Font(color="FFFFFF", bold=True)
        ws.append(["字段", "值"])
        for cell in ws[1]:
            cell.fill = blue
            cell.font = white_bold
        summary = run.summary
        fields = [
            ("Calculation ID", run.id),
            ("Status", run.status.value),
            ("Manifest SHA-256", run.manifest_hash),
            ("Snapshot SHA-256", snapshot.manifest_hash),
            ("Functional unit", summary.functional_unit),
            ("Boundary", summary.boundary),
            ("Impact method", summary.impact_method),
            ("Total kg CO2e", float(summary.total_kg_co2e)),
            ("Data quality", summary.data_quality_status),
            ("Submitted by", run.submitted_by),
            ("Approved by", run.approved_by),
        ]
        for row in fields:
            ws.append(row)
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 70

        stage = wb.create_sheet("阶段贡献")
        stage.append(["阶段", "kg CO2e"])
        for cell in stage[1]:
            cell.fill = blue
            cell.font = white_bold
        for key in ("raw_materials", "inbound_transport", "manufacturing", "packaging"):
            stage.append([key, float(getattr(summary, key))])

        iso = wb.create_sheet("ISO14067分类")
        iso.append(["类别", "kg CO2e"])
        for cell in iso[1]:
            cell.fill = blue
            cell.font = white_bold
        for key in ("aircraft", "biogenic_emissions", "biogenic_removals", "fossil", "land_use_change"):
            iso.append([key, float(getattr(summary, key))])
        iso.append(["total", float(summary.total_kg_co2e)])

        contrib = wb.create_sheet("贡献明细")
        contrib.append(["排名", "维度", "编码", "名称", "kg CO2e", "元数据"])
        for cell in contrib[1]:
            cell.fill = blue
            cell.font = white_bold
        for item in sorted(run.contributions, key=lambda x: x.rank or 999999):
            contrib.append(
                [
                    item.rank,
                    item.dimension,
                    item.code,
                    item.name,
                    float(item.amount_kg_co2e),
                    str(item.metadata_json),
                ]
            )

        bom = wb.create_sheet("BOM与因子")
        bom.append(
            [
                "行号",
                "物料编码",
                "部件",
                "阶段",
                "质量kg",
                "活动量",
                "因子单位",
                "因子值",
                "因子来源",
                "openLCA process UUID",
            ]
        )
        for cell in bom[1]:
            cell.fill = blue
            cell.font = white_bold
        for line in snapshot.payload["bom"]:
            bom.append(
                [
                    line["line_no"],
                    line["material_code"],
                    line["part_name"],
                    line["stage"],
                    float(Decimal(line["mass_kg"])),
                    float(Decimal(line["activity_amount"])),
                    line["factor_activity_unit"],
                    float(Decimal(line["factor_value"])),
                    line["factor_source"],
                    (line.get("mapping") or {}).get("process_uuid"),
                ]
            )

        energy = wb.create_sheet("制造能耗")
        energy.append(["工序编码", "工序名称", "活动量", "单位", "因子编码", "因子值", "因子来源"])
        for cell in energy[1]:
            cell.fill = blue
            cell.font = white_bold
        for item in snapshot.payload.get("energy", []):
            energy.append(
                [
                    item.get("process_code"),
                    item.get("name"),
                    float(Decimal(item["amount"])),
                    item.get("unit"),
                    item.get("factor_code"),
                    float(Decimal(item["factor_value"])),
                    item.get("source"),
                ]
            )

        transport = wb.create_sheet("入厂运输")
        transport.append(
            ["物料编码", "运输方式", "距离km", "质量kg", "装载率", "因子编码", "因子值", "数据来源"]
        )
        for cell in transport[1]:
            cell.fill = blue
            cell.font = white_bold
        for item in snapshot.payload.get("transport", []):
            transport.append(
                [
                    item.get("material_code"),
                    item.get("mode"),
                    float(Decimal(item["distance_km"])),
                    float(Decimal(item["mass_kg"])),
                    float(Decimal(item["load_factor"])) if item.get("load_factor") is not None else None,
                    item.get("factor_code"),
                    float(Decimal(item["factor_value"])),
                    item.get("source"),
                ]
            )

        exceptions = wb.create_sheet("异常项")
        exceptions.append(["来源", "代码", "说明"])
        for cell in exceptions[1]:
            cell.fill = blue
            cell.font = white_bold
        validation_errors = snapshot.validation_errors or []
        if not validation_errors and not run.error:
            exceptions.append(["snapshot", "NONE", "无阻断异常"])
        for item in validation_errors:
            exceptions.append(["snapshot", item.get("code"), str(item)])
        if run.error:
            exceptions.append(["calculation", "CALCULATION_ERROR", run.error])

        manifest = wb.create_sheet("计算清单")
        manifest.append(["字段", "值"])
        for cell in manifest[1]:
            cell.fill = blue
            cell.font = white_bold
        manifest_rows = [
            ("Calculation ID", run.id),
            ("Calculation manifest SHA-256", run.manifest_hash),
            ("Snapshot manifest SHA-256", snapshot.manifest_hash),
            ("Snapshot version", snapshot.version),
            ("Factor set version", snapshot.factor_set_version),
            ("Route version", snapshot.payload.get("route_version")),
            ("Model template version ID", run.model_template_version_id),
            ("Impact method", run.impact_method),
            ("Engine", run.engine),
            ("Engine version", run.engine_version),
            ("Raw result object key", run.raw_result_object_key),
            ("Requested by", run.requested_by),
            ("Submitted by", run.submitted_by),
            ("Approved by", run.approved_by),
        ]
        for row in manifest_rows:
            manifest.append(row)

        audit = wb.create_sheet("审批与审计")
        audit.append(["时间", "操作人", "动作", "详情"])
        for cell in audit[1]:
            cell.fill = blue
            cell.font = white_bold
        events = self.db.scalars(
            select(AuditEvent)
            .where(AuditEvent.object_id == run.id)
            .order_by(AuditEvent.occurred_at)
        )
        for event in events:
            audit.append([event.occurred_at.isoformat(), event.actor, event.action, str(event.details)])

        for sheet in wb.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
