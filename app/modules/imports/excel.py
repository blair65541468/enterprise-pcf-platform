import io
from collections import Counter
from typing import Any

from openpyxl import load_workbook

from app.modules.imports.contracts import ImportBatch, ImportFile
from app.utils import sha256_bytes


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


class ExcelImportAdapter:
    """Parses XLSX bytes without performing validation or database writes."""

    def parse(self, files: list[tuple[str, bytes]]) -> ImportBatch:
        parsed: list[ImportFile] = []
        for file_name, data in files:
            worksheet = load_workbook(io.BytesIO(data), read_only=True, data_only=True).active
            values = list(worksheet.iter_rows(values_only=True))
            if values:
                raw_headers = [_text(value) for value in values[0]]
                counts: Counter[str] = Counter()
                headers: list[str] = []
                for header in raw_headers:
                    counts[header] += 1
                    headers.append(
                        header if counts[header] == 1 else f"{header}__{counts[header]}"
                    )
                rows = tuple(
                    dict(zip(headers, row, strict=False))
                    for row in values[1:]
                    if any(value is not None for value in row)
                )
            else:
                headers = []
                rows = ()
            parsed.append(
                ImportFile(
                    name=file_name,
                    sha256=sha256_bytes(data),
                    content=data,
                    headers=tuple(headers),
                    rows=rows,
                )
            )
        return ImportBatch(source="excel", files=tuple(parsed))
