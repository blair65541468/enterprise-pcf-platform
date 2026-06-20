from openpyxl import load_workbook

from tests.conftest import auth
from tests.helpers import seed_valid_pilot


def test_end_to_end_calculation_approval_and_exports(client, db):
    seed_valid_pilot(db)
    snapshot_response = client.post(
        "/v1/products/INT-WD-001/snapshots",
        json={
            "factor_set_version": "2026.06",
            "route_version": "WD-ROUTE-V1",
            "boundary": "cradle_to_gate_with_packaging",
        },
        headers=auth("data-owner", "data_submitter"),
    )
    assert snapshot_response.status_code == 200, snapshot_response.text
    snapshot = snapshot_response.json()
    assert snapshot["validation_errors"] == []

    calc_response = client.post(
        "/v1/calculations",
        json={
            "sku": "INT-WD-001",
            "snapshot_version": snapshot["version"],
            "model_template_version": "WARDROBE-GATE-V1",
            "factor_set_version": "2026.06",
            "route_version": "WD-ROUTE-V1",
            "impact_method": "IPCC-2021-ISO14067-GWP100",
            "boundary": "cradle_to_gate_with_packaging",
            "idempotency_key": "INT-WD-001-TEST-V1",
        },
        headers=auth("data-owner", "data_submitter"),
    )
    assert calc_response.status_code == 200, calc_response.text
    run = calc_response.json()
    assert run["status"] == "calculated"
    assert float(run["summary"]["total_kg_co2e"]) == 7.525
    assert float(run["summary"]["raw_materials"]) == 6
    assert float(run["summary"]["packaging"]) == 1
    assert float(run["summary"]["manufacturing"]) == 0.5
    assert float(run["summary"]["inbound_transport"]) == 0.025

    run_id = run["id"]
    assert client.get(f"/v1/calculations/{run_id}/export.json").status_code == 403

    submit = client.post(
        f"/v1/calculations/{run_id}/submit",
        headers=auth("data-owner", "data_submitter"),
    )
    assert submit.status_code == 200

    same_user = client.post(
        f"/v1/calculations/{run_id}/approve",
        headers=auth("data-owner", "lca_reviewer"),
    )
    assert same_user.status_code == 409

    approve = client.post(
        f"/v1/calculations/{run_id}/approve",
        headers=auth("lca-reviewer", "lca_reviewer"),
    )
    assert approve.status_code == 200
    assert approve.json()["status"] == "approved"

    json_export = client.get(f"/v1/calculations/{run_id}/export.json")
    assert json_export.status_code == 200
    assert json_export.json()["manifest_hash"]
    excel_export = client.get(f"/v1/calculations/{run_id}/export.xlsx")
    assert excel_export.status_code == 200
    from io import BytesIO

    workbook = load_workbook(BytesIO(excel_export.content), read_only=True)
    assert {
        "摘要",
        "阶段贡献",
        "ISO14067分类",
        "贡献明细",
        "BOM与因子",
        "制造能耗",
        "入厂运输",
        "异常项",
        "计算清单",
        "审批与审计",
    } <= set(workbook.sheetnames)

    events = client.get(f"/v1/calculations/{run_id}/audit-events").json()
    assert [e["action"] for e in events] == [
        "calculation.queued",
        "calculation.completed",
        "calculation.submitted",
        "calculation.approved",
    ]


def test_idempotency_returns_same_run(client, db):
    seed_valid_pilot(db)
    snapshot = client.post(
        "/v1/products/INT-WD-001/snapshots",
        json={},
        headers=auth("owner", "data_submitter"),
    ).json()
    request = {
        "sku": "INT-WD-001",
        "snapshot_version": snapshot["version"],
        "model_template_version": "WARDROBE-GATE-V1",
        "factor_set_version": "2026.06",
        "route_version": "WD-ROUTE-V1",
        "impact_method": "IPCC-2021-ISO14067-GWP100",
        "boundary": "cradle_to_gate_with_packaging",
        "idempotency_key": "same-key",
    }
    first = client.post("/v1/calculations", json=request, headers=auth("owner", "data_submitter"))
    second = client.post("/v1/calculations", json=request, headers=auth("owner", "data_submitter"))
    assert first.json()["id"] == second.json()["id"]


def test_calculation_rejects_route_version_different_from_snapshot(client, db):
    seed_valid_pilot(db)
    snapshot = client.post(
        "/v1/products/INT-WD-001/snapshots",
        json={},
        headers=auth("owner", "data_submitter"),
    ).json()
    response = client.post(
        "/v1/calculations",
        json={
            "sku": "INT-WD-001",
            "snapshot_version": snapshot["version"],
            "model_template_version": "WARDROBE-GATE-V1",
            "factor_set_version": "2026.06",
            "route_version": "UNFROZEN-ROUTE",
            "impact_method": "IPCC-2021-ISO14067-GWP100",
            "boundary": "cradle_to_gate_with_packaging",
            "idempotency_key": "route-mismatch",
        },
        headers=auth("owner", "data_submitter"),
    )
    assert response.status_code == 409
